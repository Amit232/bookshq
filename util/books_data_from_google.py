from openpyxl import Workbook
import sys
from collections import defaultdict
from apiclient.discovery import build

books_metadata_dictionary = defaultdict(dict)
# wb = load_workbook('Metadata_134.xlsx')
wb = Workbook()
sheet = wb.create_sheet(title="Books_metadata_2", index=0)
api_key = "AIzaSyAPb8ehPUYe6q4qyLL0LUVz_3wJu4nmR7E"
service = build('books', 'v1', developerKey=api_key)


def get_book_title_from_txt_file(path_to_file):
    """
    This method is used to read book titles into a list from the txt file provided as command line argument
    """
    book_title_list = []
    with open(path_to_file) as file_handler:
        for line in file_handler:
            line = line.strip()
            book_title_list.append(line)
    print book_title_list
    return book_title_list


def get_author_name(data):
    try:
        return data["items"][0]["volumeInfo"]["authors"][0].encode(
            'ascii', 'ignore')
    except:
        return "Not Available"


def get_book_description(data):
    try:
        return data["items"][0]["volumeInfo"]["description"].encode(
            'ascii', 'ignore')
    except:
        return "Not Available"


def get_isbn(data):
    return data["items"][0]["volumeInfo"]["industryIdentifiers"][0][
                "identifier"].encode('ascii', 'ignore')


def get_category(data):
    try:
        return data["items"][0]["volumeInfo"]["categories"][0]
    except:
        return "Not Available"


def get_image_url(data):
    try:
        return data["items"][0]["volumeInfo"]["imageLinks"][
                "thumbnail"]
    except:
        return "Not Available"


def get_average_rating(data):
    try:
        return data["items"][0]["volumeInfo"]["averageRating"]
    except:
        return "Not Available"


def get_page_count(data):
    try:
        return data["items"][0]["volumeInfo"]["pageCount"]
    except:
        return "Not Available"


def get_book_metadata():
    # Dictionary with book title as key and related information such as author, isbn, category, description etc..
    book_title_list = get_book_title_from_txt_file(sys.argv[1])
    failed_books_list = []
    for book_title in book_title_list:
        print "Book title is " + book_title
        request = service.volumes().list(source='public', q=book_title)
        response = request.execute()
        data = response
        print type(data)
        print "JSON data loaded"
        try:
            books_metadata_dictionary[book_title]["author"] = get_author_name(data)
            books_metadata_dictionary[book_title]["description"] = get_book_description(data)
            books_metadata_dictionary[book_title]["isbn"] = get_isbn(data)
            books_metadata_dictionary[book_title]["category"] = get_category(data)
            books_metadata_dictionary[book_title]["image_url"] = get_image_url(data)
            books_metadata_dictionary[book_title]["average_rating"] = get_average_rating(data)
            books_metadata_dictionary[book_title]["pageCount"] = get_page_count(data)

        except Exception as e:
            print "Exception occured ", e
            print "Failed to get metadata for " + book_title
            print "Putting this book into a failed list"
            failed_books_list.append(book_title)
            continue
    print failed_books_list
    return books_metadata_dictionary


def dump_metadata_to_excel_file():
    global sheet
    global wb
    sheet["A1"] = "Title"
    sheet["B1"] = "Author"
    sheet["C1"] = "Description"
    sheet["D1"] = "Category"
    sheet["E1"] = "ISBN"
    sheet["F1"] = "Image Link"
    sheet["G1"] = "Average Rating"
    sheet["H1"] = "Page Count"

    final_dictionary = get_book_metadata()
    print final_dictionary.keys()
    print sheet.max_row
    for row, key in zip(range(2, 900), final_dictionary.keys()):
        try:
            sheet["A" + str(row)] = key
            sheet["B" + str(row)] = final_dictionary[key]["author"]
            sheet["C" + str(row)] = final_dictionary[key]["description"]
            sheet["D" + str(row)] = final_dictionary[key]["category"]
            sheet["E" + str(row)] = final_dictionary[key]["isbn"]
            sheet["F" + str(row)] = final_dictionary[key]["image_url"]
            sheet["G" + str(row)] = final_dictionary[key]["average_rating"]
            sheet["H" + str(row)] = final_dictionary[key]["pageCount"]
        except Exception as e:
            print("Exception ", e , " occured for " + key)
            continue

    wb.save("Metadata.xlsx")
dump_metadata_to_excel_file()
