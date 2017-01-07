from openpyxl import Workbook
from PIL import Image
from openpyxl import load_workbook
import sys
import urllib, json, cStringIO

wb = load_workbook('Metadata_134.xlsx')
sheet = wb.get_sheet_by_name("Books_metadata")


def get_thumbnail_link_from_workbook():
    my_dict = {}
    for row in range(2, sheet.max_row):
        title = sheet["A" + str(row)].value
        image_link = sheet["F" + str(row)].value
        my_dict[title] = image_link

    return my_dict


def get_image_from_url():
    count = 0
    final_dict = get_thumbnail_link_from_workbook()
    for key, value in zip(final_dict.keys(), final_dict.values()):
        # print "The key is %s and the corresponding value is %s" % (key, value)
        if value is not None:
            image_link = cStringIO.StringIO(urllib.urlopen(value).read())
            img = Image.open(image_link)
            img.save(key + ".png")
            count += 1
        else:
            print "The title %s doesnt have image_link" %(key)
    print "The total number of books we have got image for is", count

get_image_from_url()
