from openpyxl import load_workbook
import requests

wb = load_workbook('Metadata_initial.xlsx')
sheet = wb.get_sheet_by_name("Books_metadata_2")
# sheet = wb.get_sheet_by_name("Sheet")
not_available_string = "Not Available"
exception_occurred_string = "Exception Occured for this"


def get_thumbnail_link_from_workbook():
    my_dict = {}
    for row in range(2, sheet.max_row):
        title = sheet["A" + str(row)].value
        image_link = sheet["F" + str(row)].value
        # The image link field is not opulated for some books
        # Hence the below lines of code
        if image_link != not_available_string and image_link != "" and image_link != exception_occurred_string:
            my_dict[title] = image_link
        else:
            with open("not_available_image_links.txt", "a") as not_available_file_handler:
                not_available_file_handler.write(title)
                not_available_file_handler.write("\n")

    print "THE LENGTH OF THE DICTONARY IS ", len(my_dict)
    return my_dict


def get_image_from_url():
    count = 0
    final_dict = get_thumbnail_link_from_workbook()
    for key, value in zip(final_dict.keys(), final_dict.values()):
        # print "The key is %s and the corresponding value is %s" % (key, value)
        if value is not None:
            excel_sheet_title = key
            splitted = excel_sheet_title.split()
            pre_destination_file_name = ''.join(str(elem) for elem in splitted)
            destination_file_name = ''.join(e for e in pre_destination_file_name if e.isalnum()) + ".png"
            print "THE DESTINATION FILE NAME IS ", destination_file_name

            image_file = requests.get(value)
            with open(destination_file_name, "wb") as destination_file_handler:
                destination_file_handler.write(image_file.content)
            count += 1
        else:
            print "The title %s doesnt have image_link" %(key)
    print "The total number of books we have got image for is", count

get_image_from_url()
